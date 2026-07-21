"""
Minimal test: convert catalog JSON to Shopee mass upload xlsx, excluding existing products.

Usage: python scripts/test-convert-catalog-to-xlsx.py
Output: docs/data/product-catalog-upload_YYYY-MM-DD.xlsx
"""

import json
from datetime import date
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent

CATALOG_PATH = BASE_DIR / "docs" / "data" / "product-catalog-tw.json"
EXISTING_PATH = BASE_DIR / "docs" / "data" / "sample-existing-products.json"
TEMPLATE_PATH = BASE_DIR / "docs" / "shopee-official-tool" / "Shopee_mass_upload_2026-07-18_basic_template_fixed.xlsx"
OUTPUT_PATH = BASE_DIR / "docs" / "data" / f"product-catalog-upload_{date.today().isoformat()}.xlsx"

DATA_SHEET = "上傳模板"
DATA_START_ROW = 7

FIELD_MAP = {
    "ps_category": "ps_category",
    "ps_product_name": "ps_product_name",
    "ps_product_description": "ps_product_description",
    "ps_sku_parent_short": "ps_sku_short",
    "ps_price": "ps_price",
    "ps_stock": "ps_stock",
    "ps_length": "ps_length",
    "ps_width": "ps_width",
    "ps_height": "ps_height",
    "ps_item_cover_image": "ps_item_cover_image",
    "ps_item_image_1": "ps_item_image_1",
    "ps_item_image_2": "ps_item_image_2",
    "ps_item_image_3": "ps_item_image_3",
    "ps_item_image_4": "ps_item_image_4",
    "ps_item_image_5": "ps_item_image_5",
    "ps_item_image_6": "ps_item_image_6",
    "ps_item_image_7": "ps_item_image_7",
    "ps_item_image_8": "ps_item_image_8",
    "ps_brand": "ps_brand",
    "ps_weight": "ps_weight",
}

DEFAULTS = {
    "ps_hs_code": "49019900",
    "ps_tax_code": "GEN_Zero",
    "ps_weight": "0.5",
}


def load_catalog():
    with open(CATALOG_PATH, encoding="utf-8") as f:
        return json.load(f)


def load_existing_skus():
    with open(EXISTING_PATH, encoding="utf-8") as f:
        items = json.load(f)
    return {item["sku"] for item in items if item.get("sku") and item["sku"] != "-"}


def filter_new_products(catalog, existing_skus):
    new_products = []
    skipped_skus = []
    for p in catalog:
        sku = p.get("ps_sku_short", "")
        if sku and sku in existing_skus:
            skipped_skus.append(sku)
        else:
            new_products.append(p)
    return new_products, skipped_skus


def build_col_index(ws):
    """Return {api_field_name: column_index} from header row."""
    col_index = {}
    for cell in ws[1]:
        if cell.value and isinstance(cell.value, str):
            name = cell.value.split("|")[0]
            col_index[name] = cell.column - 1
    return col_index


def write_output(products, template_path, output_path, col_index):
    wb = openpyxl.load_workbook(template_path)
    ws = wb[DATA_SHEET]

    for row_idx, product in enumerate(products):
        row_num = DATA_START_ROW + row_idx
        for api_field, json_field in FIELD_MAP.items():
            col = col_index.get(api_field)
            if col is None:
                continue
            raw = product.get(json_field)
            val = str(raw) if raw is not None and raw != "" else DEFAULTS.get(api_field, "")
            ws.cell(row=row_num, column=col + 1, value=val)

    wb.save(output_path)
    wb.close()
    return output_path


def main():
    print("=== Load catalog ===")
    catalog = load_catalog()
    print(f"  Catalog products: {len(catalog)}")

    print("\n=== Load existing SKUs (simulated from seller.shopee.tw) ===")
    existing_skus = load_existing_skus()
    print(f"  Existing SKUs: {len(existing_skus)}")
    for sku in sorted(existing_skus):
        print(f"    - {sku}")

    print("\n=== Filter out existing products ===")
    new_products, skipped = filter_new_products(catalog, existing_skus)
    print(f"  Skipped: {len(skipped)} items")
    for s in skipped:
        print(f"    - {s}")
    print(f"  To output: {len(new_products)} items")

    print("\n=== Write to Excel ===")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[DATA_SHEET]
    col_index = build_col_index(ws)
    wb.close()

    if OUTPUT_PATH.exists():
        OUTPUT_PATH.unlink()
    output_path = write_output(new_products, TEMPLATE_PATH, OUTPUT_PATH, col_index)

    print(f"\nOutput file: {output_path}")
    print(f"  Size: {output_path.stat().st_size / 1024:.1f} KB")

    print("\n=== Verify ===")
    size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"  File size: {size_mb:.2f} MB [{'PASS' if size_mb <= 3.0 else 'FAIL'}]")
    print(f"  Format: .xlsx [PASS]")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[DATA_SHEET]
    data_rows = sum(1 for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True) if any(v is not None for v in row))
    wb.close()
    print(f"  Data rows: {data_rows} [{'PASS' if data_rows == len(new_products) else 'FAIL'}]")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[DATA_SHEET]
    first = {}
    for cell in ws[DATA_START_ROW]:
        header = ws.cell(row=1, column=cell.column).value
        if header:
            name = header.split("|")[0] if isinstance(header, str) else header
            first[name] = cell.value
    wb.close()
    print(f"  First row:")
    for k, v in first.items():
        print(f"    {k}: {v}")


if __name__ == "__main__":
    main()